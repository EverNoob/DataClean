__author__ = 'Rowbot'

import simplejson as json
import openpyxl
import xlrd


def json_to_csv(jsonfilepath, outfilepath):
   json_dict = json.load(jsonfilepath)
   outfile = open(outfilepath, mode='w')
   column_headers = list(json_dict[0].keys())
   for header in column_headers:
       outfile.write(str(header)+',')


def build_json_from_clean(filename):
    wb = xlrd.open_workbook(filename)
    ws = wb.sheet_by_index(0)
    temp_list = []
    column_titles = list(ws.row_values(0))
    for rownum in range(1, ws.nrows):
        row_values = ws.row_values(rownum)
        temp_dict = {}
        for titlenum in range(0, len(column_titles)):
            temp_dict[column_titles[titlenum]] = row_values[titlenum]
        temp_list.append(temp_dict)
    return temp_list


def build_brandcode_lookup(jsf):
    temp_dict= {}
    for item in jsf:
        temp_dict[item['Brand']] = item['Brand Code']
        temp_dict[item['Brand Code']] = item['Brand']

    return temp_dict


def main():
    j = json.dumps(build_json_from_clean('Cleaned data file/ShipData61815 - Example.xlsx'), sort_keys=True, indent=4 * ' ')
    with open('clean.json', 'w') as f:
        f.write(j)
        f.close()
    test = build_brandcode_lookup(build_json_from_clean())
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Brand Code')
    ws = wb.get_sheet_by_name('Brand Code')
    i = 1
    ws.cell(row=1, column=1).value = 'Brand Name'
    ws.cell(row=1, column=2).value = 'Brand Code'
    for key in test.keys():
        ws.cell(row=i, column=1).value = key
        ws.cell(row=i, column=2).value = test[key]
        i += 1
    wb.save('testlookup.xlsx')


if __name__ == '__main__':
    main()