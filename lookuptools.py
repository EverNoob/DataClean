__author__ = 'Rowbot'

import openpyxl
import simplejson as json


def build_brand_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File Updated.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    brand_dict = {}
    for rownum in range(2, ws.get_highest_row() + 1):
        itemid = ws.cell(row=rownum, column=9).value
        distributor = ws.cell(row=rownum, column=6).value
        if itemid not in brand_dict:
            brand_dict[itemid] = {distributor: {'Brand Code': ws.cell(row=rownum, column=2).value,
                                                'Brand': ws.cell(row=rownum, column=3).value,
                                                'Varietal Code': ws.cell(row=rownum, column=4).value,
                                                'Varietal': ws.cell(row=rownum, column=5).value,
                                                'SKU Tag': ws.cell(row=rownum, column=11).value,
                                                'Portfolio': ws.cell(row=rownum, column=24).value,
                                                'Category': ws.cell(row=rownum, column=25).value,
                                                'Sales/Key Acct Rep': ws.cell(row=rownum, column=26).value,
                                                'ISM': ws.cell(row=rownum, column=27).value,
                                                'IBM': ws.cell(row=rownum, column=28).value,
                                                'SKU Cost': ws.cell(row=rownum, column=31).value,
                                                'Sales Rep': ws.cell(row=rownum, column=8).value,
                                                'SKU DA': ws.cell(row=rownum, column=32).value,
                                                'Total DA$': ws.cell(row=rownum, column=33).value,
                                                'GP$/CASE': ws.cell(row=rownum, column=34).value,
                                                'Total GP$': ws.cell(row=rownum, column=35).value
                                                }}
        elif distributor not in brand_dict[itemid]:
            brand_dict[itemid][distributor] = {'Brand Code': ws.cell(row=rownum, column=2).value,
                                               'Brand': ws.cell(row=rownum, column=3).value,
                                               'Varietal Code': ws.cell(row=rownum, column=4).value,
                                               'Varietal': ws.cell(row=rownum, column=5).value,
                                               'SKU Tag': ws.cell(row=rownum, column=11).value,
                                               'Portfolio': ws.cell(row=rownum, column=24).value,
                                               'Category': ws.cell(row=rownum, column=25).value,
                                               'Sales/Key Acct Rep': ws.cell(row=rownum, column=26).value,
                                               'ISM': ws.cell(row=rownum, column=27).value,
                                               'IBM': ws.cell(row=rownum, column=28).value,
                                               'SKU Cost': ws.cell(row=rownum, column=31).value,
                                               'Sales Rep': ws.cell(row=rownum, column=8).value,
                                               'SKU DA': ws.cell(row=rownum, column=32).value,
                                               'Total DA$': ws.cell(row=rownum, column=33).value,
                                               'GP$/CASE': ws.cell(row=rownum, column=34).value,
                                               'Total GP$': ws.cell(row=rownum, column=35).value
                                               }
        #in the case that the item id is in the brand dict, and the distributor is within the itemid dictionary, it is a redundant entry and isn't used.

    with open('JSON Files/brandlookup.json', 'w') as f:
        f.write(json.dumps(brand_dict, sort_keys=True, indent=4 * ' '))


def build_innovation_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    innovation_dict = {}
    for rownum in range(2, ws.get_highest_row() + 1):
        # look for whether the brand in a "core brand" and not record the brand/innovation relationship
        if 'core' not in ws.cell(row=rownum, column=24).value.lower() and 'innovation' in ws.cell(row=rownum,
                                                                                                  column=8).value.lower():
            innovation_dict[ws.cell(row=rownum, column=3).value] = ws.cell(row=rownum, column=8).value
    with open('JSON Files/innovation_brands.json', 'w') as f:
        f.write(json.dumps(innovation_dict, sort_keys=True, indent=4 * ' '))


def build_state_by_distributor_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    state_by_distributor_dict = {}
    # TODO: account for the distributors that operate in multiple states,
    #  perhaps by making each dist a ref. to a list
    for rownum in range(2, ws.get_highest_row() + 1):
        state_by_distributor_dict[ws.cell(row=rownum, column=6).value] = ws.cell(row=rownum, column=7).value
    with open('JSON Files/state_by_distributor.json', 'w') as f:
        f.write(json.dumps(state_by_distributor_dict, sort_keys=True, indent=4 * ' '))


def build_region_by_state_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    region_by_state_dict = {}
    for rownum in range(2, ws.get_highest_row() + 1):
        region = ws.cell(row=rownum, column=8).value
        state = ws.cell(row=rownum, column=7).value
        if state not in region_by_state_dict:
            region_by_state_dict[state]= region
    #return region_by_state_dict
    with open('JSON Files/regionbystate.json', 'w') as f:
        f.write(json.dumps(region_by_state_dict, sort_keys=True, indent=4 * ' '))


def build_regionrep_by_region_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/Lookup Tables.xlsx')
    ws = wb.get_sheet_by_name('Region Rep')
    temp_dict = {}
    for rownum in range(2, ws.get_highest_row()+1):
        #TODO find out if the above line is including or overshooting the last row
        temp_dict[ws.cell(row=rownum, column=1).value] = ws.cell(row=rownum, column=2).value
    with open('JSON Files/regionrepbyregion.json', 'w') as f:
        f.write(json.dumps(temp_dict))

def build_varietal_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    varietal_by_code = {}
    for rownum in range(2, ws.get_highest_row() + 1):
        vc = ws.cell(row=rownum, column=4).value
        varietal = ws.cell(row=rownum, column=5).value
        if vc not in varietal_by_code:
            varietal_by_code[vc]= varietal
    #return region_by_state_dict
    with open('JSON Files/varietalbycode.json', 'w') as f:
        f.write(json.dumps(varietal_by_code, sort_keys=True, indent=4 * ' '))

def build_naluai_brands_list():
    wb = openpyxl.load_workbook('Lookup Tables/Lookup Tables.xlsx')
    ws = wb.get_sheet_by_name('Naluai Brands')
    naluai_brands = []
    for rownum in range(2, ws.get_highest_row() +1):
        brand = ws.cell(row=rownum, column=1).value
        naluai_brands.append(brand)
    with open('JSON Files/naluai_brand_list.json', 'w') as f:
        f.write(json.dumps(naluai_brands, sort_keys=True, indent=4 * " "))

def build_gangel_brands_list():
    wb = openpyxl.load_workbook('Lookup Tables/Lookup Tables.xlsx')
    ws = wb.get_sheet_by_name('Gangel Brands')
    gangel_brands = []
    for rownum in range(2, ws.get_highest_row() +1):
        brand = ws.cell(row=rownum, column=1).value
        gangel_brands.append(brand)
    with open('JSON Files/gangel_brand_list.json', 'w') as f:
        f.write(json.dumps(gangel_brands, sort_keys=True, indent=4 * " "))


def build_swift_brands_list():
    wb = openpyxl.load_workbook('Lookup Tables/Lookup Tables.xlsx')
    ws = wb.get_sheet_by_name('Swift Brands')
    swift_brands = []
    for rownum in range(2, ws.get_highest_row() +1):
        brand = ws.cell(row=rownum, column=1).value
        swift_brands.append(brand)
    with open('JSON Files/swift_brand_list.json', 'w') as f:
        f.write(json.dumps(swift_brands, sort_keys=True, indent=4 * " "))

def build_salesperson_brands_list(salesperson_name):
    wb = openpyxl.load_workbook('Lookup Tables/Lookup Tables.xlsx')
    ws = wb.get_sheet_by_name("" + salesperson_name + ' Brands')
    brands = []
    for rownum in range(2, ws.get_highest_row() +1):
        brand = ws.cell(row=rownum, column=1).value
        brands.append(brand)
    with open('JSON Files/' + salesperson_name.lower() + '_brand_list.json', 'w') as f:
        f.write(json.dumps(brands, sort_keys=True, indent=4 * " "))


def build_size_lookup():
    wb = openpyxl.load_workbook('Lookup Tables/2014 and YTD 2015 Shipment Data File.xlsx')
    ws = wb.get_active_sheet()
    size_dict = {}
    for rownum in range(2, ws.get_highest_row() +1):
        size_code = ws.cell(row=rownum, column=13).value
        size = ws.cell(row=rownum, column=11).value.split(' ')[-1]
        if size_code not in size_dict:
            size_dict[size_code] = size
    with open('JSON Files/size_by_letter_code.json', 'w') as f:
        f.write(json.dumps(size_dict, sort_keys=True, indent=4 * " "))


build_brand_lookup()



