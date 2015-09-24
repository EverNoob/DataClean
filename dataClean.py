
__author__ = 'Rowbot'

import openpyxl
import simplejson as json
import time
import datetime as dt
# import os
#
# abspath = os.path.abspath(__file__)
# dname = os.path.dirname(abspath)
# os.chdir(dname)




__libchange__ = False


REGION_SPELLING_DICT = {'MIDATLANTI': 'Mid Atlantic',
                        'NEW ENGLAN': 'New England',
                        'TRI-STATES': 'Tri-States',
                        'CENTRAL': 'Central',
                        'OH VALLEY': 'Ohio Valley',
                        'TEXAS': 'Texas Region',
                        'CAROLINAS': 'Carolinas',
                        'FLORIDA': 'Florida',
                        'GULF STATE': 'Gulf States',
                        'MOUNTAIN': 'Mountain',
                        'SOUTHWEST': 'Southwest',
                        'IN': 'International'
}
INNOVATION_EAST = ['Mid Atlantic', 'New England', 'Tri-States', 'Carolinas', 'Florida', 'Gulf States']

KEY_LIST = ['Item code w/o vintage', 'Brand Code', 'Brand', 'Varietal Code', 'Varietal', 'Distributor', 'State',
            'Sales Rep', 'Item ID', 'Item', 'SKU Tag', 'Item Pre', 'Size', 'Month', 'Year', 'Date', 'Document Type',
            'Warehouse', 'Document #', 'Opposite #', 'Sales $', 'Total Cases', 'Vintage', 'Portfolio', 'Category',
            'Sales/Key Acct Rep', 'ISM', 'IBM', 'Customer ID', 'Sales FOB', 'SKU Cost', 'SKU DA', 'Total DA$',
            'GP$/CASE', 'Total GP$']

new_items = []


def excel_date(date1):
    temp = dt.datetime(1899, 12, 30)
    delta = date1 - temp
    return float(delta.days)

def prettyprint(d):
    print(json.dumps(d, sort_keys=True, indent=4 * ' '))

def write_to_excel(filename, datalist):
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    rownum = 2
    # fill in the column header row
    for i in range(0, len(KEY_LIST)):
        ws.cell(row=1, column=i + 1).value = KEY_LIST[i]
    for item in datalist:
        try:
            for i in range(0, len(KEY_LIST)):
                ws.cell(row=rownum, column=i + 1).value = item[KEY_LIST[i]]
        except KeyError as ke:
            print("IT was himmm what caused the keyError>>> " + ws.cell(row=rownum, column=1).value)
        rownum += 1
    wb.save(filename)


class DataCleaner:

        # size_dict = json.load(open('JSON Files/size_by_letter_code.json'))
        # brand_dict = json.load(open('JSON Files/brandlookup.json', 'r'))
        # state_by_distributor_dict = json.load(open('JSON Files/state_by_distributor.json', 'r'))
        # region_by_state_dict = json.load(open('JSON Files/regionbystate.json'))
        # sales_rep_by_region = json.load(open('JSON Files/regionrepbyregion.json'))
        # varietal_by_code = json.load(open('JSON Files/varietalbycode.json'))
        # brand_manager_by_brand = json.load(open('JSON Files/brand_manager_by_brand.json', 'r'))
    def __init__(self):
        self.size_dict = json.load(open('JSON Files/size_by_letter_code.json', 'r'))
        self.brand_dict = json.load(open('JSON Files/brandlookup.json', 'r'))
        self.state_by_distributor_dict = json.load(open('JSON Files/state_by_distributor.json', 'r'))
        self.region_by_state_dict = json.load(open('JSON Files/regionbystate.json', 'r'))
        self.sales_rep_by_region = json.load(open('JSON Files/regionrepbyregion.json', 'r'))
        self.varietal_by_code = json.load(open('JSON Files/varietalbycode.json', 'r'))
        self.brand_manager_by_brand = json.load(open('JSON Files/brand_manager_by_brand.json', 'r'))
        # naluai_brand_list = json.load(open('JSON Files/naluai_brand_list.json'))
        # gangel_brand_list = json.load(open('JSON Files/gangel_brand_list.json'))
        # swift_brand_list = json.load(open('JSON Files/swift_brand_list.json'))
        # jackson_brand_list = json.load(open('JSON Files/jackson_brand_list.json'))
        # owens_brand_list = json.load(open('JSON Files/owens_brand_list.json'))
        # nelson_brand_list = json.load(open('Json Files/nelson_brand_list.json'))

    def update_brand_dict(self):
        with open('JSON Files/brandlookup.json', 'r') as f:
            self.brand_dict = json.load(f)

    def update_brand_manager_by_brand(self):
        with open('JSON Files/brand_manager_by_brand.json', 'r') as f:
            self.brand_manager_by_brand = json.load(f)

    def update_size_dict(self):
        with open('JSON Files/size_by_letter_code.json', 'r') as f:
            self.size_dict = json.load(f)

    def update_varietal_by_code(self):
        with open('JSON files/varietalbycode.json', 'r') as f:
            self.varietal_by_code = json.load(f)

    def remove_vintage(self, n):
        return n[0:-3] + n[-1]

    # reads in raw excel data and output a list of dictionary objects for each row

    def build_raw_list(self, filename):
        # Iterate through each row in worksheet and fetch values into dict
        wb = openpyxl.load_workbook(filename)
        sh = wb.get_active_sheet()
        i = 0
        temp_list = []
        row_tuples = sh.rows
        for rt in row_tuples[2:]:
            if rt[2].value != 'PAC' and "CONSUMER" not in rt[0].value and "Kirkland" not in rt[4].value and "zBarter" not in \
                    rt[0].value and rt[2].value != "" and 'wine country connect' not in rt[0].value.lower():
                data_item = {}
                i += 1
                data_item['Customer Name'] = rt[0].value
                data_item['Ship-to State'] = rt[1].value
                data_item['Salesperson Code'] = rt[2].value
                data_item['Item No.'] = rt[3].value
                data_item['Description'] = rt[4].value
                data_item['Product Group Code'] = rt[5].value
                data_item['Posting Month'] = rt[6].value
                data_item['Posting Date'] = excel_date(rt[7].value)
                data_item['Year'] = rt[7].value.year
                data_item['Document Type'] = rt[8].value
                data_item['Location Code'] = rt[9].value
                data_item['Document No.'] = rt[10].value
                data_item['Quantity'] = rt[11].value
                data_item['Sales Amount (Actual)'] = rt[12].value
                data_item['Quantity (positive)'] = rt[13].value
                data_item['Vintage'] = rt[14].value
                data_item['Customer No.'] = rt[15].value
                data_item['Brand Code'] = rt[16].value
                data_item['Varietal Code'] = rt[17].value
                temp_list.append(data_item)
        print('The number of raw items of interest is: ' + str(i))
        return temp_list


    # def fix_sales_rep_initial(obj_dict):
    #     region = obj_dict['Sales Rep']
    #     brand = obj_dict['Brand']
    #     category = obj_dict['Category']
    #     if region in sales_rep_by_region:
    #         obj_dict['Sales/Key Acct Rep'] = sales_rep_by_region[region]
    #     if brand in jackson_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Jackson'
    #     if brand in owens_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Owens'
    #     if brand in gangel_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Gangel'
    #     if brand in naluai_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Naluai'
    #     if brand in swift_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Swift'
    #     if brand in nelson_brand_list:
    #         obj_dict['Sales/Key Acct Rep'] = 'Nelson'
    #     if category == 'Total Wine' or 'total wine' in obj_dict['Distributor'].lower():
    #         obj_dict['Sales/Key Acct Rep'] = 'Bukoskey'
    #     return obj_dict


    def fix_sales_rep(self, obj_dict):
        region = obj_dict['Sales Rep']
        brand = obj_dict['Brand']
        category = obj_dict['Category']
        if region in self.sales_rep_by_region:
            obj_dict['Sales/Key Acct Rep'] = self.sales_rep_by_region[region]
        if brand in self.brand_manager_by_brand:
            obj_dict['Sales/Key Acct Rep'] = self.brand_manager_by_brand[brand]
        if category == 'Total Wine' or 'total wine' in obj_dict['Distributor'].lower():
            obj_dict['Sales/Key Acct Rep'] = 'Bukoskey'
        return obj_dict


    def generate_clean_data_list(self, rdl):
        scrubbed_data_list = []

        for raw_dict in rdl:
            clean_objdict = {}
            clean_objdict['Item code w/o vintage'] = self.remove_vintage(raw_dict['Item No.'])
            clean_objdict['Brand Code'] = raw_dict['Item No.'][:3]
            clean_objdict['Distributor'] = distributor = raw_dict['Customer Name']
            clean_objdict['State'] = raw_dict['Ship-to State']
            clean_objdict['Varietal Code'] = var = raw_dict['Item No.'][3:6]

            if var in self.varietal_by_code:
                clean_objdict['Varietal'] = self.varietal_by_code[clean_objdict['Varietal Code']]
            else:
                clean_objdict['Varietal'] = raw_dict['Varietal Code']

            if raw_dict['Ship-to State'] != '' and raw_dict['Ship-to State'] in self.region_by_state_dict:
                clean_objdict['Sales Rep'] = self.region_by_state_dict[raw_dict['Ship-to State']]
            else:
                clean_objdict['Sales Rep'] = raw_dict['Salesperson Code']

            clean_objdict['Item ID'] = itemid = raw_dict['Item No.']
            clean_objdict['Item'] = raw_dict['Description']
            if 'f8' not in clean_objdict['Item ID'].lower():
                clean_objdict['Item Pre'] = raw_dict['Item No.'][:9]
            else:
                clean_objdict['Item Pre'] = raw_dict['Item No.'][:8]
            clean_objdict['Size'] = itemid[-1]
            clean_objdict['Month'] = raw_dict['Posting Month']
            clean_objdict['Year'] = raw_dict['Year']
            clean_objdict['Date'] = raw_dict['Posting Date']
            clean_objdict['Document Type'] = raw_dict['Document Type']
            clean_objdict['Warehouse'] = raw_dict['Location Code']
            clean_objdict['Document #'] = raw_dict['Document No.']
            clean_objdict['Opposite #'] = raw_dict['Quantity']
            clean_objdict['Sales $'] = raw_dict['Sales Amount (Actual)']
            clean_objdict['Total Cases'] = raw_dict['Quantity (positive)']
            clean_objdict['Vintage'] = raw_dict['Vintage']
            clean_objdict['Customer ID'] = raw_dict['Customer No.']
            clean_objdict['Sales FOB'] = raw_dict['Sales Amount (Actual)'] / raw_dict['Quantity (positive)']
            clean_objdict['SKU DA'] = ''
            clean_objdict['Total DA$'] = ''
            clean_objdict['GP$/CASE'] = ''
            clean_objdict['Total GP$'] = ''

            # brand dicionary dependent items
            if itemid in self.brand_dict:
                if distributor in self.brand_dict[itemid]:
                    clean_objdict['Brand'] = self.brand_dict[itemid][distributor]['Brand']
                    clean_objdict['SKU Tag'] = self.brand_dict[itemid][distributor]['SKU Tag']
                    clean_objdict['Portfolio'] = self.brand_dict[itemid][distributor]['Portfolio']
                    clean_objdict['Category'] = self.brand_dict[itemid][distributor]['Category']
                    clean_objdict['Sales/Key Acct Rep'] = self.brand_dict[itemid][distributor]['Sales/Key Acct Rep']
                    clean_objdict['ISM'] = self.brand_dict[itemid][distributor]['ISM']
                    clean_objdict['IBM'] = self.brand_dict[itemid][distributor]['IBM']
                    clean_objdict['SKU Cost'] = self.brand_dict[itemid][distributor]['SKU Cost']
                    clean_objdict['Sales Rep'] = self.brand_dict[itemid][distributor]['Sales Rep']
                    # clean_objdict['SKU DA'] = self.brand_dict[itemid][distributor]['SKU DA']
                    # clean_objdict['Total DA$'] = self.brand_dict[itemid][distributor]['Total DA$']
                    # clean_objdict['GP$/CASE'] = self.brand_dict[itemid][distributor]['GP$/CASE']
                    # clean_objdict['Total GP$'] = self.brand_dict[itemid][distributor]['Total GP$']

                else:
                    # pick the first item out of the dictionary and use it's values to take
                    # advantage of the fact that they are of the same brand and vintage
                    # region, category, salesperson take figuring out
                    #NOTE: I may be able to create a lookup table for distributors name: {region: ####, salesrep: #####, etc...}
                    reference_item = self.brand_dict[itemid][list(self.brand_dict[itemid])[0]]
                    clean_objdict['Brand'] = reference_item['Brand']
                    clean_objdict['SKU Tag'] = reference_item['SKU Tag']
                    clean_objdict['Portfolio'] = reference_item['Portfolio']
                    clean_objdict['Category'] = reference_item['Category']
                    clean_objdict = self.fix_sales_rep(clean_objdict)

                    #TODO figure out a more elegent solution
                    clean_objdict['ISM'] = reference_item['ISM']
                    clean_objdict['IBM'] = reference_item['IBM']
                    clean_objdict['SKU Cost'] = reference_item['SKU Cost']
                    # clean_objdict['SKU DA'] = reference_item['SKU DA']
                    # clean_objdict['Total DA$'] = reference_item['Total DA$']
                    # clean_objdict['GP$/CASE'] = reference_item['GP$/CASE']
                    # clean_objdict['Total GP$'] = reference_item['Total GP$']
            #otherwise the itemcode is not yet in the database
            else:
                new_items.append(itemid)
                print(itemid, ": new code")
                standby_brand = "#N/A"
                continue_searching = True
                key_list = []
                for key in self.brand_dict.keys():
                    if itemid[:3] == key[:3] and continue_searching:
                        standby_brand = self.brand_dict[key][list(self.brand_dict[key])[0]]['Brand']
                        continue_searching = False
                    if itemid[:7] in key:
                        #collect a list of keys that match the lineitem's brand
                        key_list.append(key)
                #if this is a new brand...
                if len(key_list) == 0:
                    clean_objdict['Brand'] = standby_brand
                    if standby_brand == '#N/A':
                        clean_objdict['SKU Tag'] = '#N/A'
                    else:
                        clean_objdict['SKU Tag'] = standby_brand + clean_objdict['Varietal'] + clean_objdict['Size']
                    clean_objdict['Portfolio'] = '#N/A'
                    clean_objdict['Category'] = '#N/A'
                    clean_objdict = self.fix_sales_rep(clean_objdict)
                    clean_objdict['ISM'] = '#N/A'
                    clean_objdict['IBM'] = '#N/A'
                    clean_objdict['SKU Cost'] = '#N/A'
                    # clean_objdict['SKU DA'] = '#N/A'
                    # clean_objdict['Total DA$'] = '#N/A'
                    # clean_objdict['GP$/CASE'] = '#N/A'
                    # clean_objdict['Total GP$'] = '#N/A'
                else:
                    suitable_data_donor_found = False
                    suitable_donor = None
                    for key in key_list:
                        #if a distributor relationship is found for this same brand, use that info to fill in blanks
                        if distributor in self.brand_dict[key]:
                            suitable_data_donor_found = True
                            suitable_donor = self.brand_dict[key][distributor]
                            break
                    try:
                        if suitable_donor != None:
                            clean_objdict['Brand'] = suitable_donor['Brand']
                            clean_objdict['SKU Tag'] = suitable_donor['SKU Tag']
                            clean_objdict['Portfolio'] = suitable_donor['Portfolio']
                            clean_objdict['Category'] = suitable_donor['Category']
                            clean_objdict['Sales/Key Acct Rep'] = suitable_donor['Sales/Key Acct Rep']
                            clean_objdict['ISM'] = suitable_donor['ISM']
                            clean_objdict['IBM'] = suitable_donor['IBM']
                            clean_objdict['SKU Cost'] = suitable_donor['SKU Cost']
                            clean_objdict = self.fix_sales_rep(clean_objdict)
                            # clean_objdict['SKU DA'] = suitable_donor['SKU DA']
                            # clean_objdict['Total DA$'] = suitable_donor['Total DA$']
                            # clean_objdict['GP$/CASE'] = suitable_donor['GP$/CASE']
                            # clean_objdict['Total GP$'] = suitable_donor['Total GP$']

                        else:
                            reference_item = self.brand_dict[key][list(self.brand_dict[key])[0]]
                            clean_objdict['Brand'] = brand = reference_item['Brand']
                            clean_objdict['SKU Tag'] = brand + " " + clean_objdict['Varietal'] + " " + self.size_dict[clean_objdict['Size']]
                            clean_objdict['Portfolio'] = '#N/A'
                            clean_objdict['Category'] = '#N/A'
                            clean_objdict = self.fix_sales_rep(clean_objdict)
                            clean_objdict['ISM'] = '#N/A'
                            clean_objdict['IBM'] = '#N/A'
                            clean_objdict['SKU Cost'] = '#N/A'
                            # clean_objdict['SKU DA'] = '#N/A'
                            # clean_objdict['Total DA$'] = '#N/A'
                            # clean_objdict['GP$/CASE'] = '#N/A'
                            # clean_objdict['Total GP$'] = '#N/A'
                    except KeyError as ke:
                        print('HES the ONE YOU WANt, SEIZE HIM>>' + clean_objdict['Item ID'])

            #Initiate data scrub
            clean_objdict = self.refine_item_data(clean_objdict)
            if 'Sales/Key Acct Rep' not in clean_objdict:
                clean_objdict = self.fix_sales_rep(clean_objdict)
                if 'Sales/Key Acct Rep' not in clean_objdict:
                    if raw_dict['Salesperson Code'] in REGION_SPELLING_DICT:
                        clean_objdict['Sales/Key Acct Rep'] = self.sales_rep_by_region[
                            REGION_SPELLING_DICT[raw_dict['Salesperson Code']]]
                    else:
                        state = raw_dict['Ship-to State']
                        if state is not "":
                            clean_objdict['Sales/Key Acct Rep'] = self.sales_rep_by_region[
                            self.region_by_state_dict[raw_dict['Ship-to State']]]
                        else:
                            clean_objdict['Sales/Key Acct Rep'] = "#N/A"

            scrubbed_data_list.append(clean_objdict)
        i = 0
        nalist = []
        for object in scrubbed_data_list:
            if '#N/A' in list(object.values()):
                i += 1
                nalist.append(object['Item ID'])
        print('i: ', i)
        print(nalist)
        return scrubbed_data_list

    def refine_item_data(self, obj_dict):
        if 'Portfolio' in obj_dict:
            pf = obj_dict['Portfolio'].lower()
        else:
            pf = ""
            print("Item did no have a portfolio field")
            prettyprint(obj_dict)

        # switch monterey area items to 'Southwest' region
        if 'monterey' in obj_dict['Distributor'].lower():
            obj_dict['Sales Rep'] = 'Southwest'

        # fix canada region - Step 11
        if 'canada' in obj_dict['Sales Rep'].lower():
            obj_dict['State'] = 'Canada'
            if obj_dict['Sales Rep'].lower()[0] is 'e':
                obj_dict['Sales Rep'] = 'East Canada'
            else:
                obj_dict['Sales Rep'] = 'West Canada'

        # fix international obj_dict['Sales Rep'].lower() - Step 12
        if obj_dict['Sales Rep'].lower() == 'in':
            obj_dict['Sales Rep'] = 'International'
            # cover the instance where there is a new distributor
            if obj_dict['Distributor'] in self.state_by_distributor_dict:
                obj_dict['State'] = self.state_by_distributor_dict[obj_dict['Distributor']]

        # fix AA region - Step 13
        if 'alaska airlines' in obj_dict['Distributor'].lower():
            obj_dict['Sales Rep'] = 'Airlines'

        # fix precept house regions - Step 14
        # if the item is in the 'PH' region...
        if 'ph' in obj_dict['Sales Rep'].lower():
            # and it is Core or V&E
            if 'core' in pf or 'v&e' in pf:
                # and it is a direct shipment account...
                if 'sales shipment' in obj_dict['Document Type'].lower():
                    # make the region 'Precept House'
                    obj_dict['Sales Rep'] = 'Precept House'
                # elif obj_dict['Brand'] in innovation_by_brand_dict:
                #     obj_dict['Sales Rep'] = 'Innovation'
                else:
                    # otherwise make it the correct region according to the state
                    obj_dict['Sales Rep'] = self.region_by_state_dict[obj_dict['State']]

        # fix TOTAL WINE region - Step 15
        if 'total wine' in obj_dict['Sales Rep'].lower():
            obj_dict['Sales Rep'] = 'Total Wine'

        # fix region spelling - Step 16
        if obj_dict['Sales Rep'] in REGION_SPELLING_DICT:
            obj_dict['Sales Rep'] = REGION_SPELLING_DICT[obj_dict['Sales Rep']]

        # fix Glazer's distributor title - Step 18
        if 'Glazer' in obj_dict['Distributor'] and 'Texas' in obj_dict['Distributor']:
            obj_dict['Distributor'] = 'Glazer\'s of Texas'

        # fix Odom Name and Region - Step 19
        if 'Odom Corporation - Alaska' in obj_dict['Distributor']:
            obj_dict['State'] = 'Alaska'
            obj_dict['Sales Rep'] = "Mountain"
        if 'Odom Corporation - Cour D\'Alen' in obj_dict['Distributor'] or 'Odom Corporation - Lewiston' in obj_dict[
            'Distributor']:
            obj_dict['State'] = 'ID'
            obj_dict['Sales Rep'] = 'NW WA'
        # fix HWB - Step 31
        if 'house wine box' in obj_dict['SKU Tag'].lower():
            obj_dict['Brand'] = 'House Wine Box'
            obj_dict['Category'] = '3L BIB'
        if 'house wind lone star' in obj_dict['SKU Tag'].lower():
            obj_dict['Brand'] = 'House Wine Lone Star'
            obj_dict['Category'] = 'Innovation'
        if 'ste chapelle box' in obj_dict['SKU Tag'].lower():
            obj_dict['Brand'] = 'Ste Chappelle Box'
            obj_dict['Category'] = '3L BIB'
        if 'wtso' in obj_dict['SKU Tag'].lower():
            obj_dict['Brand'] = 'WTSO'
            obj_dict['Category'] = 'Innovation'
            # Fix Grape and grain portfolio items - Step 36
        if 'grape & grain' in pf:
            if 'total wine' not in obj_dict['Sales Rep'].lower() and 'airlines' not in obj_dict[
                'Sales Rep'].lower() and 'precept house' not in obj_dict['Sales Rep'].lower():
                if obj_dict['Sales Rep'] in INNOVATION_EAST:
                    obj_dict['Sales Rep'] = 'Innovation East'
                else:
                    obj_dict['Sales Rep'] = 'Innovation West'
        if 'airlines' in obj_dict['Sales Rep'].lower():
            obj_dict['Category'] = 'Alaska Airlines'
        if 'canada' in obj_dict['Sales Rep'].lower():
            obj_dict['Category'] = 'Canada'
        if 'international' in obj_dict['Sales Rep'].lower():
            obj_dict['Category'] = 'International'
        if 'total wine' in obj_dict['Sales Rep'].lower():
            if 'red knot' in obj_dict['Brand'].lower():
                obj_dict['Brand'] = 'Red Knot TWM'
            if 'shingleback' in obj_dict['Brand'].lower():
                obj_dict['Brand'] = 'Shingleback TWM'
            if 'apex' in obj_dict['Brand'].lower():
                obj_dict['Brand'] = 'Apex TWM'
        if 'grocery outlet' in obj_dict['Distributor'].lower():
            obj_dict['Category'] = 'Closeout'
        if 'gruet' in obj_dict['Brand'].lower():
            obj_dict['Category'] = 'Gruet'
        if 'dsv' in obj_dict['Brand'].lower():
            obj_dict['Category'] = 'Gruet'
        if 'closeout' in pf:
            obj_dict['Category'] = 'Closeout'
        if 'core' in pf:
            obj_dict['Category'] = 'Core'
        if 'v&e' in pf:
            obj_dict['Category'] = 'V&E'
        if 'alaska airlines' in obj_dict['Distributor'].lower():
            if 'canoe ridge' in obj_dict['Brand'].lower():
                obj_dict['Brand'] = 'Canoe Ridge Exploration'
        return obj_dict


def main():
    start_time = time.time()
    fname = 'Input/Item Ledger Precept - 7.21.15.xlsx'
    dc = DataCleaner()
    raw_data_list = dc.build_raw_list(fname)
    clean_data = dc.generate_clean_data_list(raw_data_list)
    with open('JSON Files/data.json', 'w') as f:
        f.write(json.dumps(clean_data, sort_keys=True, indent=4 * ' '))
    write_to_excel('Cleaned Files/megatest18.xlsx', clean_data)
    print("--- %s seconds ---" % (time.time() - start_time))


if __name__ == '__main__':
    main()